from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook


SUPPORTED_CONTAINERS = {"csv", "xlsx", "ofx", "qfx", "qif"}


@dataclass
class ParsedFile:
    headers: list[str]
    rows: list[dict[str, Any]]


def detect_container(filename: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext not in SUPPORTED_CONTAINERS:
        raise ValueError(f"Unsupported file container: {ext}")
    return ext


def parse_csv(content: bytes) -> ParsedFile:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    if sample.strip():
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
    else:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = [header.strip() for header in (reader.fieldnames or [])]
    rows = [{(k or "").strip(): (v or "").strip() for k, v in row.items()} for row in reader]
    return ParsedFile(headers=headers, rows=rows)


def parse_xlsx(content: bytes, sheet_name: str | None = None) -> ParsedFile:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    iterator = worksheet.iter_rows(values_only=True)
    header_row = next(iterator, None)
    if header_row is None:
        return ParsedFile(headers=[], rows=[])

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    rows: list[dict[str, Any]] = []
    for row in iterator:
        values = ["" if cell is None else str(cell).strip() for cell in row]
        rows.append({headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers))})
    return ParsedFile(headers=headers, rows=rows)


def parse_ofx_qfx(content: bytes) -> ParsedFile:
    text = content.decode("utf-8", errors="replace")
    tx_blocks = re.findall(r"<STMTTRN>(.*?)</STMTTRN>", text, flags=re.DOTALL | re.IGNORECASE)
    rows: list[dict[str, Any]] = []

    def extract(block: str, tag: str) -> str:
        match = re.search(rf"<{tag}>([^<\r\n]+)", block, flags=re.IGNORECASE)
        return match.group(1).strip() if match else ""

    for block in tx_blocks:
        rows.append(
            {
                "posting_date": extract(block, "DTPOSTED"),
                "amount": extract(block, "TRNAMT"),
                "description": extract(block, "MEMO") or extract(block, "NAME"),
                "reference": extract(block, "FITID"),
            }
        )

    return ParsedFile(headers=["posting_date", "amount", "description", "reference"], rows=rows)


def parse_qif(content: bytes) -> ParsedFile:
    text = content.decode("utf-8", errors="replace")
    records = text.split("^\n")
    rows: list[dict[str, Any]] = []

    for record in records:
        row: dict[str, str] = {}
        for line in record.splitlines():
            if not line:
                continue
            prefix, value = line[0], line[1:].strip()
            if prefix == "D":
                row["posting_date"] = value
            elif prefix == "T":
                row["amount"] = value
            elif prefix == "P":
                row["description"] = value
            elif prefix == "M":
                row["memo"] = value
            elif prefix == "N":
                row["reference"] = value
        if row:
            rows.append(row)

    headers = sorted({k for row in rows for k in row.keys()})
    return ParsedFile(headers=headers, rows=rows)


def parse_container(content: bytes, container: str, *, sheet_name: str | None = None) -> ParsedFile:
    if container == "csv":
        return parse_csv(content)
    if container == "xlsx":
        return parse_xlsx(content, sheet_name=sheet_name)
    if container in {"ofx", "qfx"}:
        return parse_ofx_qfx(content)
    if container == "qif":
        return parse_qif(content)
    raise ValueError(f"Unsupported container: {container}")
